from bs4 import BeautifulSoup
import requests
from openpyxl import workbook, load_workbook

from selenium import webdriver
import time

driver = webdriver.Chrome()

m=1028
for page_number in range(0, m):
    driver.get(f'https://www.collegesimply.com/colleges/rank/colleges/lowest-graduation-rate/?page={page_number}')
    html_text = driver.page_source
    time.sleep(2)
    page_code = BeautifulSoup(html_text, 'lxml')
    code = page_code.find('div', class_= 'container py-5')
    table = code.find('div', class_='col')
    versities_list = table.find_all('tr', itemprop="itemListElement")  #find_all
    for versity in versities_list:

        try:
            file_path = r'file path'
            wb = load_workbook(file_path)
            ws = wb['Sheet1']
            char_3 = 'D'
            link_element = versity.find('a', itemprop = 'url')
            versity_link = link_element['href']
            driver.get(f'https://www.collegesimply.com/{versity_link}')
            html_text_2 = driver.page_source
            time.sleep(1)

            soup = BeautifulSoup(html_text_2, 'lxml')
            code = soup.find('div', class_ = 'col col-md-9 col-lg-10')
            table_data = code.find('div', class_ ='card shadow')
            enroll = table_data.find('div', class_ = 'row')
            enrollment = enroll.find('div', class_ = 'h2 mb-4').text
            ws[char_3 + str(m)].value = enrollment
            m=m+1
            wb.save(file_path)
        except:
            pass